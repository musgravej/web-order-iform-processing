import openpyxl
import re
import os
import xlrd
import csv
import datetime


def process_context_item(fle):
    print("Processing Context Item file:", fle)
    updated_file = "IForms_ContextItem_Update_{}.xlsx".format(
            datetime.datetime.strftime(datetime.date.today(), "%Y%m%d"))

    if os.path.exists(updated_file):
        os.remove(updated_file)

    oexcel = os.path.join(os.curdir, 'update_files', 'IForms_ContextItem_Update.xlsx')

    wb_template = openpyxl.load_workbook(filename=oexcel)
    ws_template = wb_template.active

    wb = xlrd.open_workbook(fle)
    active_sheet = wb.sheets()[0]
    for n, row in enumerate(range(active_sheet.nrows), 9):
        if n > 9:
            for col in range(active_sheet.ncols):
                ws_template.cell(row=n-1, column=col+1, value=active_sheet.cell(row, col).value)
                # print(active_sheet.cell(row, col).value)

    wb_template.save(updated_file)


def process_list_item(fle):
    print("Processing List Item file:", fle)
    updated_file = "IForms_ListItem_Update_{}.xlsx".format(datetime.datetime.strftime(datetime.date.today(), "%Y%m%d"))
    if os.path.exists(updated_file):
        os.remove(updated_file)

    oexcel = os.path.join(os.curdir, 'update_files', 'IForms_ListItem_Update.xlsx')

    wb = openpyxl.load_workbook(filename=oexcel)
    ws = wb.active

    with open(fle, 'r') as f:
        csvr = csv.reader(f, delimiter="\t")
        for n, line in enumerate(csvr, 9):
            for field in range(len(line)):
                ws.cell(row=n, column=field+1, value=line[field])

    wb.save(updated_file)


def process_function_map(fle):
    print("Processing Function Map file:", fle)
    updated_file = "IForms_FunctionMap_Update_{}.xlsx".format(
            datetime.datetime.strftime(datetime.date.today(), "%Y%m%d"))
    if os.path.exists(updated_file):
        os.remove(updated_file)

    oexcel = os.path.join(os.curdir, 'update_files', 'IForms_FunctionMap_Update.xlsx')

    wb_template = openpyxl.load_workbook(filename=oexcel)
    ws_template = wb_template.active

    wb = xlrd.open_workbook(fle)
    active_sheet = wb.sheets()[0]
    for n, row in enumerate(range(active_sheet.nrows), 9):
        if n > 9:
            for col in range(active_sheet.ncols):
                ws_template.cell(row=n - 1, column=col + 1, value=active_sheet.cell(row, col).value)
                # print(active_sheet.cell(row, col).value)

    wb_template.save(updated_file)


def main():
    list_item_srch = re.compile("[\s\S]*_listitem[\s\S]*.txt", re.IGNORECASE)
    context_item_srch = re.compile("[\s\S]*_contextitem[\s\S]*.xls", re.IGNORECASE)
    function_map_srch = re.compile("[\s\S]*_functionmap[\s\S]*.xls", re.IGNORECASE)

    dir_list = os.listdir(os.curdir)

    for fle in dir_list:
        if list_item_srch.match(fle):
            process_list_item(fle)

        if context_item_srch.match(fle):
            process_context_item(fle)

        if function_map_srch.match(fle):
            process_function_map(fle)

if __name__ == '__main__':
    main()
